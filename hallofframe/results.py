"""Append race results to a local .xlsx results log (spec §6.8 extension).

A finished race's results block (title / header / data rows) is appended to a
workbook whose path is configurable and typically lives in a folder a desktop
sync client mirrors to Google Drive. The app itself does no networking (spec N3);
the sync client owns upload. Writes are append-only and atomic (temp file +
``os.replace``) so a sync client never uploads a half-written file.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

from .export import data_rows
from .storage import Storage


def has_block_for_name(path, name: str) -> bool:
    """True if *name* appears as an exact column-A value in the workbook.

    Used to decide the ``(re-recorded — an earlier entry exists)`` note at save
    time. Returns False if the file is missing or openpyxl is unavailable.
    """
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return False
    p = Path(path)
    if not p.exists():
        return False
    target = (name or "").strip()
    if not target:
        return False
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                if row and row[0] is not None and str(row[0]).strip() == target:
                    return True
            return False
        finally:
            wb.close()
    except Exception:
        return False


def _last_data_row(ws) -> int:
    """Highest row index holding a non-empty value (styling does not count)."""
    last = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                if cell.row > last:
                    last = cell.row
    return last


def append_race(storage: Storage, race_id: int, out_path) -> Path | None:
    """Append one race's results block to *out_path*; return the path.

    Guards: the race must have ``ended_at`` set (raises ``ValueError``) and
    openpyxl must be importable (logs-and-no-ops to ``None`` instead of raising,
    mirroring ``races.py``). The title is bold; if the race *name* already has a
    block, the new title carries a re-recorded note. Writes atomically.
    """
    try:
        import openpyxl
    except Exception:
        return None
    from openpyxl.styles import Font

    race = storage.get_race(race_id)
    if race is None or race["ended_at"] is None:
        raise ValueError(f"race {race_id} has not ended; refusing to append")

    title = (race["name"] or "").strip()
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.exists():
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        start = _last_data_row(ws) + 2  # one blank, unstyled spacer row between blocks
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        start = 1

    if has_block_for_name(out_path, title):
        title = f"{title} (re-recorded — an earlier entry exists)"

    rows = list(data_rows(storage, race_id))  # [header, data, ...]
    ws.cell(row=start, column=1, value=title).font = Font(bold=True)
    for col, value in enumerate(rows[0], start=1):
        ws.cell(row=start + 1, column=col, value=value)
    for i, row in enumerate(rows[1:], start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=start + i, column=col, value=value)

    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(out_path.parent), suffix=".xlsx.tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(out_path))
    except Exception:
        if tmp_path is not None:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        raise
    finally:
        wb.close()
    return out_path
