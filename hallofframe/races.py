"""Race-name roster loaded from an Excel workbook (spec: one race per row).

The operator maintains an .xlsx where column A holds the race names, one per
row. The main window shows these in a dropdown and passes the selected name to
``start_race``. Reads via openpyxl; a missing file or missing library degrades
to an empty list (the UI falls back to a timestamp name). ``write_example``
creates a starter workbook if none exists.
"""
from __future__ import annotations

from pathlib import Path

HEADER = "race_name"

_EXAMPLE = [
    "Heat 1 - Men's Single",
    "Heat 2 - Men's Single",
    "Heat 3 - Men's Single",
    "Heat 4 - Women's Single",
    "Final A - Men's Double",
    "Final B - Women's Quad",
]


def _names_from_workbook(excel_path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        names: list[str] = []
        seen: set[str] = set()
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            text = str(row[0]).strip()
            if not text or text.lower() == HEADER:
                continue
            if text in seen:
                continue
            seen.add(text)
            names.append(text)
        return names
    finally:
        wb.close()


def load_race_names(excel_path) -> list[str]:
    """Return the race names in column A of the first sheet, in order."""
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return []
    if not Path(excel_path).exists():
        return []
    try:
        return _names_from_workbook(excel_path)
    except Exception:
        return []


def write_example(excel_path, names: list[str] | None = None) -> Path:
    """Write a starter workbook with a header row and example race names."""
    import openpyxl
    names = list(names) if names is not None else list(_EXAMPLE)
    p = Path(excel_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "races"
    ws.append([HEADER])
    for n in names:
        ws.append([n])
    wb.save(str(p))
    return p
