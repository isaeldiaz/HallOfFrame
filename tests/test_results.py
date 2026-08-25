"""T-results — append race results to results.xlsx (see results.md)."""
import tempfile
import unittest
from pathlib import Path

import openpyxl

from hallofframe.results import append_race, has_block_for_name
from hallofframe.storage import Storage


def make_race(storage, name, ended=True):
    rid = storage.create_race(name, 1000.0, 1000.0, "direct", 0.0, 0.0, "water", 30)
    if ended:
        storage.mark_race_ended(rid, 2000.0)
    return rid


def dump(path) -> list[list]:
    """Read a workbook into a row list, preserving blank rows."""
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return rows


def last_data_row(rows) -> int:
    """1-based index of the last row with any non-empty value."""
    for i in range(len(rows) - 1, -1, -1):
        if any(v is not None and str(v).strip() for v in rows[i]):
            return i + 1
    return 0


class TestResults(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.data_root = Path(self.tmp.name)
        self.storage = Storage(self.data_root)
        self.out = self.data_root / "results.xlsx"

    def tearDown(self):
        self.storage.close()
        self.tmp.cleanup()

    def _save(self, rid):
        return append_race(self.storage, rid, self.out)

    def test_single_block_layout(self):
        rid = make_race(self.storage, "Heat 1 - Men's Single")
        self.storage.insert_capture(rid, 1, 2000.0, 2000.0, 372.483, 0.0,
                                    bow_number="4", image_flag="approximate")
        self.storage.insert_capture(rid, 2, 3000.0, 3000.0, 373.0, 0.0,
                                    bow_number="7")
        self._save(rid)
        rows = dump(self.out)
        self.assertEqual(rows[0][0], "Heat 1 - Men's Single")
        # title is bold
        wb = openpyxl.load_workbook(str(self.out))
        self.assertTrue(wb.active.cell(row=1, column=1).font.bold)
        wb.close()
        self.assertEqual(rows[1], ["sequence", "bow_number", "elapsed_seconds",
                                   "elapsed_formatted", "wall_clock_utc",
                                   "image_file", "image_flag", "notes"])
        self.assertEqual(rows[2][:4], [1, "4", "372.483000", "6:12.483"])
        self.assertEqual(rows[3][:3], [2, "7", "373.000000"])
        # compact block: data ends the sheet (the blank spacer is implicit
        # between blocks, since openpyxl drops a fully-empty trailing row)
        self.assertEqual(last_data_row(rows), 4)
        self.assertEqual(len(rows), 4)

    def test_two_blocks_one_blank_row_no_gap_growth(self):
        r1 = make_race(self.storage, "Heat 1")
        r2 = make_race(self.storage, "Heat 2")
        self.storage.insert_capture(r1, 1, 2000.0, 2000.0, 1.0, 0.0, bow_number="1")
        self.storage.insert_capture(r2, 1, 2000.0, 2000.0, 2.0, 0.0, bow_number="2")
        self._save(r1)
        rows = dump(self.out)
        end1 = last_data_row(rows)
        self._save(r2)
        rows = dump(self.out)
        end2 = last_data_row(rows)
        # block 2's title sits one blank spacer row below block 1's last data row
        self.assertEqual(rows[end1 + 1][0], "Heat 2")
        self.assertTrue(all(v is None for v in rows[end1]))
        # no growing gap: block 2's extent (title..last data) matches block 1's
        self.assertEqual(end2 - (end1 + 2) + 1, end1)

    def test_re_recorded_note_and_prior_block_untouched(self):
        r1 = make_race(self.storage, "Heat 1 - Men's Single")
        self.storage.insert_capture(r1, 1, 2000.0, 2000.0, 1.0, 0.0, bow_number="4")
        self._save(r1)
        before = dump(self.out)
        r2 = make_race(self.storage, "Heat 1 - Men's Single")
        self.storage.insert_capture(r2, 1, 2000.0, 2000.0, 2.0, 0.0, bow_number="7")
        self._save(r2)
        rows = dump(self.out)
        noted = [r[0] for r in rows
                 if isinstance(r[0], str) and r[0].startswith(
                     "Heat 1 - Men's Single (re-recorded")]
        self.assertEqual(len(noted), 1)
        self.assertIn("an earlier entry exists", noted[0])
        # the prior block is byte-for-byte untouched
        self.assertEqual(rows[: last_data_row(before)], before[: last_data_row(before)])

    def test_soft_deleted_excluded_and_edited_bow_reflected(self):
        rid = make_race(self.storage, "Heat 1")
        c1 = self.storage.insert_capture(rid, 1, 2000.0, 2000.0, 1.0, 0.0,
                                         bow_number="4")
        self.storage.insert_capture(rid, 2, 3000.0, 3000.0, 2.0, 0.0, bow_number="7")
        self.storage.update_capture(c1, deleted=1)
        self.storage.update_capture(c1, bow_number="99")  # deleted row edited
        self._save(rid)
        rows = dump(self.out)
        self.assertEqual(rows[2][:2], [2, "7"])

    def test_unended_race_not_appended(self):
        rid = make_race(self.storage, "Heat 1", ended=False)
        with self.assertRaises(ValueError):
            self._save(rid)
        self.assertFalse(self.out.exists())

    def test_missing_openpyxl_degrades_to_noop(self):
        rid = make_race(self.storage, "Heat 1")
        import builtins
        real_import = builtins.__import__
        def _blocked(name, *a, **k):
            if name == "openpyxl":
                raise ImportError("blocked for test")
            return real_import(name, *a, **k)
        builtins.__import__ = _blocked
        try:
            result = self._save(rid)
        finally:
            builtins.__import__ = real_import
        self.assertIsNone(result)
        self.assertFalse(self.out.exists())

    def test_atomic_write_no_temp_left_and_no_truncation_on_failure(self):
        rid = make_race(self.storage, "Heat 1")
        self.storage.insert_capture(rid, 1, 2000.0, 2000.0, 1.0, 0.0, bow_number="4")
        self._save(rid)
        original = dump(self.out)

        # simulate a mid-write failure by making os.replace raise
        import hallofframe.results as results
        real_replace = results.os.replace
        results.os.replace = lambda *a, **k: (_ for _ in ()).throw(OSError("boom"))
        try:
            with self.assertRaises(OSError):
                self._save(rid)
        finally:
            results.os.replace = real_replace
        # the original workbook is intact and no temp file lingers
        self.assertEqual(dump(self.out), original)
        leftovers = list(self.data_root.glob("*.tmp")) + list(
            self.data_root.glob("*.xlsx.tmp"))
        self.assertEqual(leftovers, [])

    def test_has_block_for_name(self):
        self.assertFalse(has_block_for_name(self.out, "Heat 1"))
        rid = make_race(self.storage, "Heat 1")
        self._save(rid)
        self.assertTrue(has_block_for_name(self.out, "Heat 1"))
        self.assertFalse(has_block_for_name(self.out, "Heat 2"))
        self.assertFalse(has_block_for_name(self.data_root / "nope.xlsx", "Heat 1"))


if __name__ == "__main__":
    unittest.main()
